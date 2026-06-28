from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))
from pipeline_common import GENERATED_DIR, REPORTS_DIR, as_str, ensure_dirs, load_sources, make_id, normalize_text, write_json


def import_master_dictionary() -> dict:
    ensure_dirs()
    source = next((item for item in load_sources() if item.get("parser_type") == "dictionary"), None)
    if not source:
        raise RuntimeError("Dictionary source is not configured in data/source_files.json")

    path = Path(source["source_file"])
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [as_str(v) for v in rows[0]]
    index = {name: i for i, name in enumerate(header)}

    services = []
    synonyms = []
    seen_ids = set()
    for row_number, row in enumerate(rows[1:], start=2):
        name = as_str(row[index.get("Name_ru", 3)] if len(row) > 3 else "")
        if not name:
            continue
        specialty = as_str(row[index.get("Специальность", 1)] if len(row) > 1 else "")
        code = as_str(row[index.get("Code", 2)] if len(row) > 2 else "")
        source_id = as_str(row[index.get("ID", 0)] if len(row) > 0 else "")
        tarificator_code = as_str(row[index.get("TarificatrCode", 4)] if len(row) > 4 else "")
        service_id = make_id("svc", tarificator_code or source_id, code, name)
        if service_id in seen_ids:
            service_id = make_id(service_id, row_number)
        seen_ids.add(service_id)
        normalized = normalize_text(name)
        service = {
            "id": service_id,
            "source_dictionary_id": source_id,
            "specialty": specialty,
            "code": code,
            "name_ru": name,
            "name": name,
            "category": specialty or "Без категории",
            "description": f"{specialty}: {name}" if specialty else name,
            "tarificator_code": tarificator_code,
            "normalized_name": normalized,
            "synonyms_base": sorted(set([name, normalized, tarificator_code, code]) - {""}),
            "synonyms": sorted(set([name, normalized, tarificator_code, code]) - {""}),
        }
        services.append(service)
        for synonym in service["synonyms"]:
            synonyms.append({"service_id": service_id, "synonym": synonym, "normalized_synonym": normalize_text(synonym), "source": "master_dictionary"})

    write_json(GENERATED_DIR / "services.json", services)
    write_json(GENERATED_DIR / "service_synonyms.json", synonyms)
    report = {
        "source_file": str(path),
        "sheet": ws.title,
        "rows_seen": max(0, len(rows) - 1),
        "services_written": len(services),
        "synonyms_written": len(synonyms),
        "columns": header,
    }
    write_json(REPORTS_DIR / "dictionary_report.json", report)
    return report


if __name__ == "__main__":
    print(import_master_dictionary())
