from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))
from pipeline_common import RAW_DIR, REPORTS_DIR, as_str, build_raw_record, clean_price, ensure_dirs, load_sources, write_json


EXCEL_CONFIGS: dict[str, list[dict[str, Any]]] = {
    "clinic_6": [
        {
            "sheet": "приложение",
            "start_row": 14,
            "columns": {
                "service_code": 2,
                "raw_service_name": 3,
                "unit": 4,
                "base_without_vat": 6,
                "resident": 7,
                "non_resident_cis": 8,
                "non_resident_far": 9,
            },
        }
    ],
    "clinic_8": [
        {
            "sheet": "Страховой",
            "start_row": 21,
            "columns": {
                "raw_service_name": 2,
                "service_code": 3,
                "unit": 4,
                "insurance": 5,
            },
        },
        {
            "sheet": "востребовынные страховых",
            "start_row": 23,
            "columns": {
                "raw_service_name": 2,
                "service_code": 3,
                "unit": 4,
                "resident": 5,
            },
        },
    ],
}


def cell(row: tuple[Any, ...], one_based_index: int | None) -> Any:
    if not one_based_index:
        return None
    idx = one_based_index - 1
    return row[idx] if idx < len(row) else None


def parse_xlsx_source(source: dict) -> tuple[list[dict], dict]:
    path = Path(source["source_file"])
    clinic_id = source.get("clinic_id")
    configs = EXCEL_CONFIGS.get(clinic_id, [])
    records: list[dict] = []
    warnings: list[str] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for config in configs:
            sheet_name = config["sheet"]
            if sheet_name not in wb.sheetnames:
                warnings.append(f"Missing sheet: {sheet_name}")
                continue
            ws = wb[sheet_name]
            columns = config["columns"]
            for row_number, row in enumerate(ws.iter_rows(min_row=config["start_row"], values_only=True), start=config["start_row"]):
                raw_name = as_str(cell(row, columns.get("raw_service_name")))
                if not raw_name or raw_name.lower().startswith(("раздел", "итого")):
                    continue
                service_code = cell(row, columns.get("service_code"))
                unit = cell(row, columns.get("unit"))
                raw_row = {str(i + 1): as_str(v) for i, v in enumerate(row)}
                price_columns = [key for key in columns if key not in {"service_code", "raw_service_name", "unit"}]
                for price_type in price_columns:
                    price = clean_price(cell(row, columns[price_type]))
                    rec = build_raw_record(
                        source,
                        service_code=service_code,
                        raw_service_name=raw_name,
                        unit=unit,
                        price_kzt=price,
                        price_type=price_type,
                        source_sheet=sheet_name,
                        raw_row_json={"row_number": row_number, "cells": raw_row},
                        parser_confidence=0.96 if price else 0.3,
                        row_index=row_number,
                    )
                    if rec:
                        records.append(rec)
    except Exception as exc:
        return [], {"source_file": str(path), "parser": "excel", "status": "failed", "records": 0, "warnings": [str(exc)]}

    return records, {
        "source_file": str(path),
        "clinic_id": clinic_id,
        "parser": "excel",
        "status": "success" if records else "warning",
        "records": len(records),
        "warnings": warnings[:100],
    }


def parse_legacy_xls_source(source: dict) -> tuple[list[dict], dict]:
    path = Path(source["source_file"])
    try:
        import xlrd  # type: ignore  # noqa: F401
    except Exception:
        return [], {
            "source_file": str(path),
            "clinic_id": source.get("clinic_id"),
            "parser": "excel_legacy",
            "status": "warning",
            "records": 0,
            "warnings": ["Legacy .xls requires xlrd or a working LibreOffice conversion fallback. File was reported, not skipped silently."],
        }
    return [], {
        "source_file": str(path),
        "clinic_id": source.get("clinic_id"),
        "parser": "excel_legacy",
        "status": "warning",
        "records": 0,
        "warnings": ["xlrd is available but legacy parser implementation is intentionally deferred behind higher quality XLSX/PDF extraction."],
    }


def parse_all_excel() -> dict:
    ensure_dirs()
    all_records: list[dict] = []
    reports: list[dict] = []
    for source in load_sources():
        if source.get("parser_type") == "excel":
            records, report = parse_xlsx_source(source)
        elif source.get("parser_type") == "excel_legacy":
            records, report = parse_legacy_xls_source(source)
        else:
            continue
        all_records.extend(records)
        reports.append(report)
    write_json(RAW_DIR / "excel_price_records.json", all_records)
    write_json(REPORTS_DIR / "excel_import_report.json", {"files": reports, "records": len(all_records)})
    return {"records": len(all_records), "files": reports}


if __name__ == "__main__":
    print(parse_all_excel())
